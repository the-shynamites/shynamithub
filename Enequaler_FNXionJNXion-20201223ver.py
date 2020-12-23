#Enequaler_FNXionJNXion-20201223ver.py

# coding:utf-8

import time
import pprint
import openpyxl

from openpyxl import load_workbook

# xlsxファイル名を入力してEnterで確定する。
print("PLZ choose your XLSX file, and press Enter-key, e.g. fnxion_dequal (without [.xlsx]).")
XLSX0 = input()
XLSX1 = XLSX0 + '.xlsx'
# ワークブック(例:「fnxion_dequal.xlsx」)を開く(Excelファイルを読み込む)。
book1 = openpyxl.load_workbook(XLSX1) #formula

# sheet名を入力してEnterで確定する。
print("PLZ choose your xlsx SHEET, and press Enter-key, e.g. fnxiontest.")
SHEET1 = input()
#シート(例:「fnxiontest」)を指定する
sheet1 = book1[SHEET1]

#数式の入ったセルの範囲をR1C1表記で指定する。→各セルの冒頭に一文字だけ加える(数式にするべく冒頭にイコールを加えて、計算処理を有効化する)。

print("PLZ choose min-Row (in R1C1 style), and press Enter-key. e.g. 1.")
fromminRow = int(input())

print("PLZ choose min-Column (in R1C1 style), and press Enter-key. e.g. 4, when you'd choose column D.")
fromminColumn = int(input())

print("PLZ choose Max-Row (in R1C1 style), and press Enter-key. e.g. 10.")
toMaxRow = int(input())

print("PLZ choose Max-Column (in R1C1 style), and press Enter-key. e.g. 4, when you'd choose column D.")
toMaxColumn = int(input())


for cols1 in sheet1.iter_cols(min_row=fromminRow, min_col=fromminColumn, max_row=toMaxRow, max_col=toMaxColumn):
    for cell1 in cols1:
    	cell1.value = '=' + cell1.value  


# 2秒間の待機
time.sleep(2)

# 念の為、別名で保存しておく
book1.save(XLSX0 + '_JNXion_Enequaled_FNXion.xlsx')
 
    
# 読み込んでいたExcelファイルを閉じる
book1.close()

# “FNXionJNXion (Enequaler) in Python 3.8” produced by K.masamix “KIXAN” as the SHYNAMITES/La CHENAMITOJ, 2020/12/23.


