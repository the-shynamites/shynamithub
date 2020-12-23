#ValueKeeper_FNXionJNXion-20201223ver.py

# coding:utf-8

import time
import pprint
import openpyxl

from openpyxl import load_workbook

# xlsxファイル名を入力してEnterで確定する。
print("PLZ choose your XLSX file, and press Enter-key, e.g. fnxion (without [.xlsx]).")
XLSX0 = input()
XLSX1 = XLSX0 + '.xlsx'
# ワークブック(例:「fnxion.xlsx」)を開く(Excelファイルを読み込む)。
book1 = openpyxl.load_workbook(XLSX1,  data_only=True) #value

# sheet名を入力してEnterで確定する。
print("PLZ choose your xlsx SHEET, and press Enter-key, e.g. fnxiontest.")
SHEET1 = input()
#シート(例:「fnxiontest」)を指定する
sheet1 = book1[SHEET1]

#値の入ったセルの範囲をR1C1表記で指定する(例:D1〜D10までのセルを指定する)。

print("PLZ choose min-Row (in R1C1 style), and press Enter-key. e.g. 1.")
fromminRow = int(input())

print("PLZ choose min-Column (in R1C1 style), and press Enter-key. e.g. 4, when you'd choose column D.")
fromminColumn = int(input())

print("PLZ choose Max-Row (in R1C1 style), and press Enter-key. e.g. 10.")
toMaxRow = int(input())

print("PLZ choose Max-Column (in R1C1 style), and press Enter-key. e.g. 4, when you'd choose column D.")
toMaxColumn = int(input())


for cols1 in sheet1.iter_cols(min_row=fromminRow, min_col=fromminColumn, max_row=toMaxRow, max_col=toMaxColumn):
    for cell1 in cols1:
    	cell1.value
    

# 2秒間の待機
time.sleep(2)

#別の列セルに値を代入する(例:G1～G10までのセルに値を転記する)。
v = cell1.value

print("PLZ choose NEW min-Row (in R1C1 style), and press Enter-key. e.g. 1.")
fromNEWminRow = int(input())

print("PLZ choose NEW min-Column (in R1C1 style), and press Enter-key. e.g. 7, when you'd choose column G.")
fromNEWminColumn = int(input())

print("PLZ choose NEW Max-Row (in R1C1 style), and press Enter-key. e.g. 10.")
toNEWMaxRow = int(input())

print("PLZ choose NEW Max-Column (in R1C1 style), and press Enter-key. e.g. 7, when you'd choose column G.")
toNEWMaxColumn = int(input())


for cols4 in sheet1.iter_cols(min_row=fromNEWminRow, min_col=fromNEWminColumn, max_row=toNEWMaxRow, max_col=toNEWMaxColumn):
    for cell4 in cols4:
    	cell4.value = v
    	
# 2秒間の待機
time.sleep(2)


# 念の為、別名で保存しておく
book1.save(XLSX0 + '_ValueKept_FNXion_JNXion.xlsx')
 
    
# 読み込んでいたExcelファイルを閉じる
book1.close()

# “FNXionJNXion (ValueKeeper) in Python 3.8” produced by K.masamix “KIXAN” as the SHYNAMITES/La CHENAMITOJ, 2020/12/23.


