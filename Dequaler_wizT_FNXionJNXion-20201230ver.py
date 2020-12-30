#Dequaler_FNXionJNXion-20201230ver.py

# coding:utf-8

import time
import pprint
import openpyxl

from openpyxl import load_workbook

# xlsxファイル名を入力してEnterで確定する。
print("PLZ choose your XLSX file, and press Enter-key, e.g. fnxion (without [.xlsx]).")
XLSX0 = input()
XLSX1 = XLSX0 + '.xlsx'
# ワークブック(例:「fnxion.xlsx」)を開く(Excelファイルを読み込む)。
book1 = openpyxl.load_workbook(XLSX1) #formula

# sheet名を入力してEnterで確定する。
print("PLZ choose your xlsx SHEET, and press Enter-key, e.g. fnxiontest.")
SHEET1 = input()
#シート(例:「fnxiontest」)を指定する
sheet1 = book1[SHEET1]

#数式の入ったセルの範囲をR1C1表記で指定する。→各セルの冒頭の一文字を別の文字に替える(数式なので冒頭のイコールを除いて、計算処理を無効化する)。

print("PLZ choose min-Row (in R1C1 style), and press Enter-key. e.g. 1.")
fromminRow = int(input())

print("PLZ choose min-Column (in R1C1 style), and press Enter-key. e.g. 4, when you'd choose column D.")
fromminColumn = int(input())

print("PLZ choose Max-Row (in R1C1 style), and press Enter-key. e.g. 10.")
toMaxRow = int(input())

print("PLZ choose Max-Column (in R1C1 style), and press Enter-key. e.g. 4, when you'd choose column D.")
toMaxColumn = int(input())


for cols1 in sheet1.iter_cols(min_row=fromminRow, min_col=fromminColumn, max_row=toMaxRow, max_col=toMaxColumn):
    for cell1 in cols1:
    	cell1.value = cell1.value.replace('=', 'T', 1)


# 2秒間の待機
time.sleep(2)

# 念の為、別名で保存しておく
book1.save(XLSX0 + '_FNXion_wizT_Dequaled_JNXion.xlsx')
 
    
# 読み込んでいたExcelファイルを閉じる
book1.close()

# “FNXionJNXion (Dequaler) in Python 3.8” produced by K.masamix “KIXAN” as the SHYNAMITES/La CHENAMITOJ, 2020/12/30.


