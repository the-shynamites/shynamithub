#PivoTexlsellent-20210104ver.py 20210104-

# coding:utf-8
#pip3 install xlrd==1.2.0

import time
import pprint
import openpyxl
import xlrd
import xlwt
from openpyxl import load_workbook

import pandas as pd

#Sample of Texlsellent
df1 = pd.read_excel('ダミーテクセレント.xlsx', sheet_name = 'concat')

#列kk→Surface1 of 《AB》
piv1= df1.pivot_table(index=['kk','項番'], values=['dd','textjoin','viva'],aggfunc='sum',margins=True, margins_name='Surface1')

#列ll→Surface2 of 《AB》
piv2= df1.pivot_table(index=['ll','項番'], values=['dd','textjoin','viva'],aggfunc='sum',margins=True, margins_name='Surface2')

#列mm→Surface3 of 《AB》
piv3= df1.pivot_table(index=['mm','項番'], values=['dd','textjoin','viva'],aggfunc='sum',margins=True, margins_name='Surface3')


#列aa→One extreme Line of 《A》and《B》
piv4= df1.pivot_table(index=['aa','項番'], values=['dd','textjoin','viva'],aggfunc='sum',margins=True, margins_name='LineAlpha')

#列bb→The other extreme Line of 《A》and《B》
piv5= df1.pivot_table(index=['bb','項番'], values=['dd','textjoin','viva'],aggfunc='sum',margins=True, margins_name='LineBeta')

#列cc→The moderate Line of 《A》and《B》
piv6= df1.pivot_table(index=['cc','項番'], values=['dd','textjoin','viva'],aggfunc='sum',margins=True, margins_name='LineGamma')


#既存のExcelファイル「ピボテクセレントSurfaces.xlsx」に新たなシートとして追加し、処理結果を追記する。


with pd.ExcelWriter('ピボテクセレントSurfaces.xlsx', mode = 'a') as writer: # 追記モード20210111
    #writer.book = openpyxl.load_workbook('ピボテクセレントSurfaces.xlsx') #←これが無くても動く20210111
    piv1.to_excel(writer, sheet_name='Surface1')
    piv2.to_excel(writer, sheet_name='Surface2')
    piv3.to_excel(writer, sheet_name='Surface3')

with pd.ExcelWriter('ピボテクセレントLines.xlsx', mode = 'a') as writer: # 追記モード20210111
    #writer.book = openpyxl.load_workbook('ピボテクセレントLines.xlsx') #←これが無くても動く20210111
    piv4.to_excel(writer, sheet_name='LineAlpha')
    piv5.to_excel(writer, sheet_name='LineBeta')
    piv6.to_excel(writer, sheet_name='LineGamma')


# “PivoTexlsellent in Python 3.8” produced by K.masamix “KIXAN” as the SHYNAMITES/La CHENAMITOJ, 2021/01/04.
